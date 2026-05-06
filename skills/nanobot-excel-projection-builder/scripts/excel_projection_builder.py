#!/usr/bin/env python3
"""excel_projection_builder.py — Generate a formatted Excel financial projection workbook."""
import argparse
import json
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
ALT_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CURR_FMT = '#,##0;(#,##0);"-"'
CURR_FMT_LARGE = '#,##0,;(#,##0,);"-"'
PCT_FMT = '0.0%;0.0%;"-"'


def _fmt_currency(ws, row, num_cols, start_col=2):
    for c in range(start_col, num_cols + 2):
        cell = ws.cell(row=row, column=c)
        if cell.value is None:
            continue
        v = cell.value
        if isinstance(v, (int, float)):
            cell.number_format = CURR_FMT_LARGE if abs(v) >= 1000000 else CURR_FMT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='right')


def _fmt_pct(ws, row, start_col=2):
    for c in range(start_col, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = PCT_FMT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='right')


def _set_col_widths(ws):
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 3, 30)


def _write_section_header(ws, label, num_cols):
    ws.append([label] + [''] * num_cols)
    row = ws.max_row
    for c in range(1, num_cols + 2):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = BORDER


def _write_total_row(ws, label, values, num_cols):
    ws.append([label] + values)
    row = ws.max_row
    for c in range(1, num_cols + 2):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
    _fmt_currency(ws, row, num_cols)


def _write_data_row(ws, label, values, num_cols, shaded=False):
    ws.append([label] + values)
    row = ws.max_row
    for c in range(1, num_cols + 2):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')
        if shaded:
            cell.fill = ALT_FILL
    _fmt_currency(ws, row, num_cols)


def _title_block(ws, company_name, title, currency):
    ws.append([f"{company_name} — {title}"])
    ws.merge_cells(f'A1:{get_column_letter(20)}1')
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='left')
    ws.append([f"Currency: {currency}"])
    ws['A2'].font = Font(italic=True, size=10, color="595959")


def _header_row(ws, years, num_cols):
    ws.append([''] + years)
    row = ws.max_row
    for c in range(1, num_cols + 2):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER


def _find_metric(data, section_name, label, num_years):
    section_map = {
        'income_statement': 'income_statement',
        'balance_sheet': 'balance_sheet',
        'cash_flow': 'cash_flow',
    }
    key = section_map.get(section_name, section_name)
    rows = data.get('projections', {}).get(key, [])
    label_lower = label.lower()

    for row_data in rows:
        row_label = str(row_data.get('line_item', row_data.get('label', ''))).lower()
        if row_label == label_lower or row_label.replace(' ', '') == label_lower.replace(' ', ''):
            vals = row_data.get('values', [])[:num_years]
            return vals
    return None


def build_summary_tab(wb, data, years):
    ws = wb.active
    ws.title = "Summary"

    _title_block(ws, data.get('company_name', 'Company'), 'Financial Projection Summary', data.get('currency', 'USD'))

    ws.append([])
    ws.append(['Key Metrics'])
    ws['A4'].font = Font(bold=True, size=11, color="1F4E79")

    summary_rows = [
        ('Revenue', 'income_statement', 'revenue', False),
        ('Gross Profit', 'income_statement', 'gross_profit', False),
        ('EBITDA', 'income_statement', 'ebitda', False),
        ('EBIT', 'income_statement', 'ebit', False),
        ('Net Income', 'income_statement', 'net_income', False),
        ('Total Assets', 'balance_sheet', 'total_assets', False),
        ('Total Liabilities', 'balance_sheet', 'totalliabilities', False),
        ('Total Equity', 'balance_sheet', 'totalequity', False),
        ('Operating Cash Flow', 'cash_flow', 'operating_cash_flow', False),
        ('Free Cash Flow', 'cash_flow', 'free_cash_flow', False),
    ]

    ws.append([])
    ws.append([''] + years)
    hdr = ws.max_row
    for c in range(1, len(years) + 2):
        cell = ws.cell(row=hdr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    for label, section, key, is_bold in summary_rows:
        values = _find_metric(data, section, key, len(years))
        if values is None:
            continue
        ws.append([label] + values)
        row = ws.max_row
        for c in range(1, len(years) + 2):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL
        _fmt_currency(ws, row, len(years))
        if is_bold:
            ws.cell(row=row, column=1).font = Font(bold=True)

    ws.append([])

    ws.append(['Growth Rates'])
    ws[f'A{ws.max_row}'].font = Font(bold=True, size=11, color="1F4E79")
    growth_rows = [
        ('Revenue Growth', 'income_statement', 'revenue'),
        ('Net Income Growth', 'income_statement', 'net_income'),
    ]
    for label, section, key in growth_rows:
        prev = _find_metric(data, section, key, len(years))
        if not prev:
            continue
        rates = []
        for i, v in enumerate(prev):
            if i == 0:
                rates.append(None)
            elif prev[i - 1] and prev[i - 1] != 0:
                rates.append((v - prev[i - 1]) / abs(prev[i - 1]))
            else:
                rates.append(None)
        ws.append([label] + rates)
        row = ws.max_row
        for c in range(1, len(years) + 2):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL
        _fmt_pct(ws, row)

    _set_col_widths(ws)
    ws.freeze_panes = 'B5'


def build_income_statement_tab(wb, data, years):
    ws = wb.create_sheet("Income Statement")

    _title_block(ws, data.get('company_name', 'Company'), 'Income Statement Projection', data.get('currency', 'USD'))
    ws.append([])
    num_cols = len(years)
    _header_row(ws, years, num_cols)

    is_rows = data.get('projections', {}).get('income_statement', [])

    known_sections = {
        'revenue': 'Revenue',
        'cost of goods sold': 'COGS',
        'cogs': 'COGS',
        'gross profit': 'GROSS_PROFIT',
        'operating expenses': 'Operating Expenses',
        'opex': 'Operating Expenses',
        'selling, general & administrative': 'SG&A',
        'sg&a': 'SG&A',
        'research and development': 'R&D',
        'r&d': 'R&D',
        'ebitda': 'EBITDA',
        'depreciation & amortization': 'D&A',
        'd&a': 'D&A',
        'ebit': 'EBIT',
        'interest': 'Interest',
        'interest expense': 'Interest',
        'earnings before tax': 'EBT',
        'ebt': 'EBT',
        'income tax': 'Taxes',
        'taxes': 'Taxes',
        'net income': 'NET_INCOME',
        'net profit': 'NET_INCOME',
    }

    def get_section_key(label):
        label_lower = label.lower().strip()
        for k, v in known_sections.items():
            if k in label_lower or label_lower in k:
                return v
        return None

    for row_data in is_rows:
        label = row_data.get('line_item', row_data.get('label', ''))
        values = row_data.get('values', [])[:num_cols]
        while len(values) < num_cols:
            values.append(None)
        is_total = row_data.get('is_total', False) or row_data.get('is_bold', False)
        sk = get_section_key(label)
        if is_total or (sk and sk in ('GROSS_PROFIT', 'EBITDA', 'EBIT', 'NET_INCOME')):
            _write_total_row(ws, label, values, num_cols)
        else:
            _write_data_row(ws, label, values, num_cols, shaded=ws.max_row % 2 == 0)

    _set_col_widths(ws)
    ws.freeze_panes = 'B5'


def build_balance_sheet_tab(wb, data, years):
    ws = wb.create_sheet("Balance Sheet")

    _title_block(ws, data.get('company_name', 'Company'), 'Balance Sheet Projection', data.get('currency', 'USD'))
    ws.append([])
    num_cols = len(years)
    _header_row(ws, years, num_cols)

    bs_rows = data.get('projections', {}).get('balance_sheet', [])

    known_totals = {'total assets', 'total liabilities', 'total equity', 'total current assets',
                    'total non-current assets', 'total current liabilities', 'total long-term liabilities'}

    for row_data in bs_rows:
        label = row_data.get('line_item', row_data.get('label', ''))
        values = row_data.get('values', [])[:num_cols]
        while len(values) < num_cols:
            values.append(None)
        is_total = row_data.get('is_total', False) or row_data.get('is_bold', False)
        l_lower = label.lower().strip()
        if is_total or l_lower in known_totals:
            _write_total_row(ws, label, values, num_cols)
        else:
            _write_data_row(ws, label, values, num_cols, shaded=ws.max_row % 2 == 0)

    _set_col_widths(ws)
    ws.freeze_panes = 'B5'


def build_cash_flow_tab(wb, data, years):
    ws = wb.create_sheet("Cash Flow")

    _title_block(ws, data.get('company_name', 'Company'), 'Cash Flow Projection', data.get('currency', 'USD'))
    ws.append([])
    num_cols = len(years)
    _header_row(ws, years, num_cols)

    cf_rows = data.get('projections', {}).get('cash_flow', [])

    known_totals = {'net cash from operating activities', 'net cash from operating',
                    'operating cash flow', 'ocf',
                    'net cash from investing activities', 'investing cash flow',
                    'net cash from financing activities', 'financing cash flow',
                    'net change in cash', 'net increase in cash',
                    'ending cash', 'cash at end of period', 'ending cash balance',
                    'beginning cash', 'cash at beginning of period', 'beginning cash balance'}

    for row_data in cf_rows:
        label = row_data.get('line_item', row_data.get('label', ''))
        values = row_data.get('values', [])[:num_cols]
        while len(values) < num_cols:
            values.append(None)
        is_total = row_data.get('is_total', False) or row_data.get('is_bold', False)
        l_lower = label.lower().strip()
        if is_total or l_lower in known_totals:
            _write_total_row(ws, label, values, num_cols)
        else:
            _write_data_row(ws, label, values, num_cols, shaded=ws.max_row % 2 == 0)

    _set_col_widths(ws)
    ws.freeze_panes = 'B5'


def build_assumptions_tab(wb, data):
    ws = wb.create_sheet("Assumptions")

    _title_block(ws, data.get('company_name', 'Company'), 'Projection Assumptions', data.get('currency', 'USD'))
    ws.append([])

    assumptions = data.get('assumptions', {})

    ws.append(['Assumption', 'Value', 'Description'])
    hdr = ws.max_row
    for c in range(1, 4):
        cell = ws.cell(row=hdr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    rows = [
        ('Revenue Growth Rate', assumptions.get('revenue_growth_rate'), 'Annual revenue growth', True),
        ('COGS % of Revenue', assumptions.get('cogs_pct_of_revenue'), 'Cost of goods sold as % of revenue', True),
        ('OpEx Growth Rate', assumptions.get('opex_growth_rate'), 'Annual operating expense growth', True),
        ('CapEx (Annual)', assumptions.get('capex_annual'), 'Annual capital expenditures', False),
        ('Depreciation %', assumptions.get('depreciation_pct'), 'Depreciation as % of revenue', True),
        ('Tax Rate', assumptions.get('tax_rate'), 'Effective tax rate', True),
        ('Working Capital Days', assumptions.get('working_capital_days'), 'Days sales outstanding / working capital cycle', False),
        ('Debt (Starting)', assumptions.get('debt'), 'Starting long-term debt', False),
        ('Equity (Starting)', assumptions.get('equity'), 'Starting shareholders equity', False),
    ]

    for i, (name, value, desc, is_pct) in enumerate(rows, start=1):
        ws.append([name, value, desc])
        row = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL
        if c == 2 and value is not None:
            if is_pct and isinstance(value, float) and 0 < value < 1:
                ws.cell(row=row, column=2).number_format = PCT_FMT
            elif isinstance(value, (int, float)):
                ws.cell(row=row, column=2).number_format = CURR_FMT_LARGE

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.freeze_panes = 'A5'


def main():
    parser = argparse.ArgumentParser(description='Build a formatted Excel financial projection workbook from JSON data.')
    parser.add_argument('input_json', help='Path to projection JSON file')
    parser.add_argument('--output', dest='output', default=None, help='Output .xlsx path')
    parser.add_argument('--company-name', dest='company_name', default=None,
                        help='Override company name in output')
    args = parser.parse_args()

    try:
        with open(args.input_json, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: File not found: {args.input_json}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON: {e}", file=sys.stderr)
        sys.exit(1)

    company = args.company_name or data.get('company_name', 'Company')
    years = data.get('projections', {}).get('years', [])

    if not years:
        print("Error: No 'projections.years' found in JSON. Cannot build workbook.", file=sys.stderr)
        sys.exit(1)

    output_path = args.output
    if not output_path:
        safe_name = company.replace(' ', '_').replace('/', '_')
        output_path = f"{safe_name}_projection.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_tab(wb, data, years)
    build_income_statement_tab(wb, data, years)
    build_balance_sheet_tab(wb, data, years)
    build_cash_flow_tab(wb, data, years)
    build_assumptions_tab(wb, data)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print("Tabs: Summary | Income Statement | Balance Sheet | Cash Flow | Assumptions")


if __name__ == '__main__':
    main()

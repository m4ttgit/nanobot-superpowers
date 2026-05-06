#!/usr/bin/env python3
import argparse
import json
import os
import sys
import re
import datetime
from pathlib import Path

def _import_optional(module_name):
    try:
        return __import__(module_name)
    except Exception:
        return None

def _try_import_pdfplumber():
    try:
        import pdfplumber
        return pdfplumber
    except Exception:
        return None

def _try_import_tabula():
    try:
        from tabula import read_pdf  # type: ignore
        return read_pdf
    except Exception:
        return None

def _parse_number(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        return None
    # Remove currency symbols and thousands separators
    s = s.replace('$', '').replace(',', '').strip()
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1].strip()
    if s == '':
        return None
    try:
        v = float(s)
        if neg:
            v = -v
        return v
    except ValueError:
        return None

def _detect_period_from_text(text: str) -> str:
    if not text:
        return ""
    # FY pattern first
    m = re.search(r'FY\s*(\d{2,4})', text, re.IGNORECASE)
    if m:
        y = m.group(1)
        if len(y) == 2:
            y = '20' + y
        return f'FY{y}'
    # Look for a 4-digit year
    years = re.findall(r'\b(?:19|20)\d{2}\b', text)
    if years:
        return years[-1]
    return ""

def _detect_statement_type(text: str) -> str:
    if not text:
        return "unknown"
    t = text.lower()
    if 'income statement' in t or 'statement of income' in t or 'profit and loss' in t:
        return 'income_statement'
    if 'balance sheet' in t or 'statement of financial position' in t:
        return 'balance_sheet'
    if 'cash flow' in t or 'statement of cash flows' in t:
        return 'cash_flow'
    return 'unknown'

def _convert_table_to_rows(table_rows, headers):
    # Convert a raw list of rows (strings) into structured rows with first column as label
    converted = []
    if not table_rows:
        return converted
    for r in table_rows:
        if not r:
            continue
        label = str(r[0]).strip() if len(r) > 0 else ''
        values = []
        for c in r[1:]:
            values.append(_parse_number(c))
        converted.append({"label": label, "values": values})
    return converted

def _ensure_dir(path: Path) -> None:
    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)

def extract_from_pdfplumber(pdf_path: str):
    import pdfplumber  # type: ignore
    statements = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = (page.extract_text() or "")
            stype = _detect_statement_type(text)
            period = _detect_period_from_text(text)

            tables = page.extract_tables()
            if not tables:
                continue
            for tbl in tables:
                if not tbl or len(tbl) < 2:
                    continue
                header = [str(x).strip() if x is not None else '' for x in tbl[0]]
                # Use the rest of the rows as data if available
                data_rows = tbl[1:]
                rows = _convert_table_to_rows(data_rows, header)
                if not rows:
                    continue
                headers = header[1:] if len(header) > 1 else []
                statements.append({
                    "type": stype or "unknown",
                    "page": i,
                    "title": {
                        'income_statement': 'Income Statement',
                        'balance_sheet': 'Balance Sheet',
                        'cash_flow': 'Cash Flow Statement',
                        'unknown': 'Statement'
                    }.get(stype, 'Statement'),
                    "period": period,
                    "headers": headers,
                    "rows": rows,
                    "notes": []
                })
    return {
        "source_file": os.path.basename(pdf_path),
        "extracted_at": datetime.datetime.utcnow().isoformat() + "Z",
        "statements": statements
    }

def extract_with_tabula(pdf_path: str):
    # Fallback extraction using tabula-py
    try:
        read_pdf = __import__('tabula').read_pdf  # type: ignore
    except Exception:
        read_pdf = None
    if read_pdf is None:
        return None
    try:
        dfs = read_pdf(pdf_path, pages='all', multiple_tables=True)
    except Exception:
        return None
    statements = []
    period = ''
    # Heuristic: try to deduce period from the first page text via file content (rare here)
    for idx, df in enumerate(dfs, start=1):
        if df is None or df.empty:
            continue
        header = [str(h) for h in df.columns.tolist()]
        # Build rows by iterating rows; assume first column is label
        rows = []
        for _, row in df.iterrows():
            vals = [row[c] for c in df.columns[1:]]
            label = str(row[df.columns[0]]) if df.columns[0] in df.columns else str(row[0])
            numeric = [_parse_number(v) for v in vals]
            rows.append({"label": label, "values": numeric})
        headers = header[1:]
        # Try to guess type by keywords in header
        t = 'unknown'
        # crude guess from header text
        if any(h and 'income' in str(h).lower() for h in header):
            t = 'income_statement'
        elif any(h and 'balance' in str(h).lower() for h in header):
            t = 'balance_sheet'
        elif any(h and 'cash' in str(h).lower() for h in header):
            t = 'cash_flow'
        statements.append({
            "type": t,
            "page": idx,
            "title": {
                'income_statement': 'Income Statement',
                'balance_sheet': 'Balance Sheet',
                'cash_flow': 'Cash Flow Statement',
                'unknown': 'Statement'
            }.get(t, 'Statement'),
            "period": period,
            "headers": headers,
            "rows": rows,
            "notes": []
        })
    return {
        "source_file": os.path.basename(pdf_path),
        "extracted_at": datetime.datetime.utcnow().isoformat() + "Z",
        "statements": statements
    }

def main():
    parser = argparse.ArgumentParser(description='Extract financial tables from a PDF and output JSON/Excel-like data.')
    parser.add_argument('pdf_path', help='Path to the input PDF file')
    parser.add_argument('--output-dir', dest='output_dir', default=None, help='Output directory (default: same as input)')
    parser.add_argument('--format', dest='fmt', choices=['json','xlsx','both'], default='json', help='Output format: json, xlsx, or both')
    parser.add_argument('--statement-type', dest='stmt_type', choices=['auto','income','balance','cashflow'], default='auto', help='Override statement type detection (auto, income, balance, cashflow)')
    args = parser.parse_args()

    pdf_path = os.path.abspath(args.pdf_path)
    if not os.path.isfile(pdf_path):
        print(f"Error: PDF file not found: {pdf_path}", file=sys.stderr)
        sys.exit(2)

    output_dir = Path(args.output_dir) if args.output_dir else Path(pdf_path).parent / (Path(pdf_path).stem + "_extracted")
    _ensure_dir(output_dir)

    # Attempt extraction with pdfplumber first, then fallback to tabula
    data = None
    pdfplumber = _try_import_pdfplumber()
    if pdfplumber is not None:
        data = extract_from_pdfplumber(pdf_path)
    else:
        # Try to load tabula and extract
        tabula_fn = _try_import_tabula()
        if tabula_fn is not None:
            data = extract_with_tabula(pdf_path)
        else:
            print("Error: Neither pdfplumber nor tabula-py could be imported. Install one of them:\n  pip install pdfplumber\nor\n  pip install tabula-py", file=sys.stderr)
            sys.exit(3)

    if data is None:
        print("Error: Failed to extract tables from the PDF. Ensure the PDF contains tabular data and try again.")
        sys.exit(4)

    # If user requested a specific statement type, we can mark unmatched ones accordingly
    if args.stmt_type != 'auto':
        # Normalize type mappings
        map_overrides = {
            'income': 'income_statement',
            'balance': 'balance_sheet',
            'cashflow': 'cash_flow'
        }
        target = map_overrides.get(args.stmt_type, 'unknown')
        for s in data.get('statements', []):
            s['type'] = target
            s['title'] = {
                'income_statement': 'Income Statement',
                'balance_sheet': 'Balance Sheet',
                'cash_flow': 'Cash Flow Statement',
                'unknown': 'Statement'
            }.get(target, 'Statement')

    base = Path(pdf_path).stem
    # JSON path
    json_path = output_dir / f"{base}_extracted.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"Wrote JSON: {json_path}")

    # Excel output using openpyxl
    if args.fmt in ('xlsx', 'both'):
        excel_path = output_dir / f"{base}_extracted.xlsx"
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            excel_path = None
            print("Warning: openpyxl not installed. Install with: pip install openpyxl\n  Excel output skipped.")

        if excel_path:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
            ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            CURR_FMT = '#,##0;(#,##0);"-"'

            thin = Side(style='thin')
            BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

            for stmt_idx, stmt in enumerate(data.get('statements', []), start=1):
                ws = wb.create_sheet(
                    title=f"{stmt['type'].replace('_',' ').title()} {stmt_idx}"
                )

                ws.merge_cells(f'A1:{get_column_letter(len(stmt.get("headers",[""]))+1)}1')
                title_cell = ws['A1']
                title_cell.value = f"{stmt['title']} — {stmt['period']} (Page {stmt['page']})"
                title_cell.font = Font(bold=True, size=12, color="1F4E79")
                title_cell.alignment = Alignment(horizontal='left')

                ws.append([])
                headers = stmt.get('headers', [])
                ws.append([''] + headers)
                hdr_row = ws.max_row
                for col_idx in range(2, len(headers) + 2):
                    cell = ws.cell(row=hdr_row, column=col_idx)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = BORDER

                for row_data in stmt.get('rows', []):
                    ws.append([row_data['label']] + row_data['values'])
                    data_row_idx = ws.max_row
                    for col_idx in range(2, len(headers) + 2):
                        cell = ws.cell(row=data_row_idx, column=col_idx)
                        val = cell.value
                        if isinstance(val, (int, float)) and val is not None:
                            cell.number_format = CURR_FMT
                        cell.border = BORDER
                        cell.alignment = Alignment(horizontal='right')
                    if data_row_idx % 2 == 0:
                        for col_idx in range(1, len(headers) + 2):
                            ws.cell(row=data_row_idx, column=col_idx).fill = ALT_FILL

                for col_cells in ws.columns:
                    length = max(len(str(cell.value) if cell.value else '') for cell in col_cells)
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 40)

                ws.freeze_panes = 'B4'

            wb.save(excel_path)
            print(f"Wrote Excel workbook: {excel_path}")

    print("Extraction complete.")

if __name__ == '__main__':
    main()
